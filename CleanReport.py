from tkinter import *
from tkinter import messagebox
from tkinter import simpledialog
from tkinter.filedialog import askopenfilename
from tkinter.filedialog import asksaveasfilename
import openpyxl 
from openpyxl import load_workbook
import math
from win32com.client import GetObject
import wmi
import os

def pickAFile():
    loc = askopenfilename()
    loc2 = "S:\ApplicationList.xlsx"
    file = load_workbook(loc)
    mod = load_workbook(loc2)
    #sheet = file.get_sheet_by_name("Report0")
    #sheet2 = mod.get_sheet_by_name("Foglio1")
    sheet = file[file.sheetnames[0]]
    sheet2 = mod[mod.sheetnames[0]]
    Confronta(sheet, sheet2, file)
    messagebox.showinfo("Alert", "Script eseguito con successo.")
    
    

def Confronta(sheet, sheet2, file):
    for x in range(1, 1000):  
        var = sheet2["A" + str(x)].internal_value
        if type(var) == type(None):
            break
        for y in range(1,  150):
            var2 = sheet["C" + str(y)].internal_value
            if type(var2) == type(None):
                break
            if str(var) in str(var2):
                sheet["C" + str(y)] = ""

    for x in range(1, 1000):  
        var = ""
        for y in range(1,  150):
            var2 = sheet["C" + str(y)].internal_value
            if type(var2) == type(None):
                break
            if var2 == var:
                sheet.delete_rows(y)
    messagebox.showinfo("Alert", "Salva il file.")
    file.save(asksaveasfilename(filetypes=[("Excel files", "*.xlsx")], defaultextension = "xlsx"))
    

def getSerialMonitor(string, w2):
    objWMI = GetObject('winmgmts:\\\\'+ string + '\\root\WMI').InstancesOf('WmiMonitorID')
    cont = 1
    temp=""
    for obj in objWMI:
        if obj.SerialNumberID != None:
            temp = "".join(str(chr(i)) for i in obj.SerialNumberID)
            cont +=1
            Label(w2, text="Monitor Serial ID: " + temp, anchor="center").pack()
            

def getPrinter(string, w2):
    asset = wmi.WMI(string)
    cont = 2
    Label(w2, text="\nPrinters: ").pack()
    for info in asset.Win32_Printer():
        temp = info.caption
        cont += 1
        Label(w2, text=temp, anchor="center").pack()

def getAssetInfo():
    loc2 = "S:\ApplicationList.xlsx"
    mod = load_workbook(loc2)
    sheet2 = mod.get_sheet_by_name("Foglio1")
    string = simpledialog.askstring("", "Inserisci il nome macchina")
    w2 = Toplevel(window)
    w2.geometry("300x400")
    scrollbar = Scrollbar(w2)
    scrollbar.pack(side=RIGHT, fill=Y)
    listbox = Listbox(w2, width=50, height=20, yscrollcommand=scrollbar.set)
    listbox.insert(END, "prova")
    listbox.pack(side=LEFT, fill=BOTH, expand=True)
    scrollbar.config(command=listbox.yview)
    getSerialMonitor(string, w2)
    getPrinter(string, w2)


window = Tk()
window.title("Rollout Info 0.0.2")
window.geometry("300x300")
btn1 = Button(window, text = "Pulizia File StockReport", command = pickAFile)
btn1.place(relx=0.5, rely=0.5, anchor=CENTER)
btn2 = Button(window, text = "Monitor e Stampanti", command = getAssetInfo)
btn2.pack(side = TOP, padx=80, pady=80)




window.mainloop()
